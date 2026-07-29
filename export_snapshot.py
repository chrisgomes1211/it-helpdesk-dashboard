#!/usr/bin/env python3
"""
Read IT_Helpdesk_Ticket_Log.xlsx (sheet "Ticket Log") and embed
the data as a JSON snapshot into index.html, replacing the array
between the __EMBEDDED_SNAPSHOT_START__ / __EMBEDDED_SNAPSHOT_END__ markers.

Usage:
    python export_snapshot.py

Requires: openpyxl (pip install openpyxl)
"""

import json, os, re, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl is required — install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

XLSX_PATH = "IT_Helpdesk_Ticket_Log.xlsx"
SHEET_NAME = "Ticket Log"
INDEX_PATH = "index.html"


def cell_value(cell):
    """Return a string representation of a cell value."""
    val = cell.value
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if val is None:
        return ""
    return val


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"Error: {XLSX_PATH} not found in the current directory.", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2))
    headers = [cell.value for cell in ws[1]]

    data = []
    for row in rows:
        ticket = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i] is not None:
                ticket[headers[i]] = cell_value(cell)
        if ticket:
            data.append(ticket)

    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    with open(INDEX_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    pattern = (
        r"(// __EMBEDDED_SNAPSHOT_START__\s*\n\s*const SNAPSHOT_DATA = )"
        r"\[.*?\]"
        r"(\s*;\s*\n\s*// __EMBEDDED_SNAPSHOT_END__)"
    )
    replacement = lambda m: m.group(1) + json_str + m.group(2)
    new_html = re.sub(pattern, replacement, html, count=1, flags=re.DOTALL)

    if new_html == html:
        print("Warning: markers not found in index.html — no changes made.", file=sys.stderr)
        sys.exit(1)

    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"Embedded {len(data)} tickets from {XLSX_PATH} into {INDEX_PATH}")


if __name__ == "__main__":
    main()
